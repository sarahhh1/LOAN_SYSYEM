import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import sqlite3
from datetime import datetime
import openpyxl


# Create or connect to the SQLite database
def create_db():
    conn1 = sqlite3.connect('students.db')
    c1 = conn1.cursor()
    c1.execute('''CREATE TABLE IF NOT EXISTS Students (
                 DateLoan TEXT,
                 DateAssumedReturn TEXT,
                 StudentName TEXT,
                 StudentID TEXT,
                 PhoneNum TEXT,
                 Email TEXT,
                 ProjectTutor TEXT,
                 GipNum TEXT,
                 SerialNum TEXT
                 )''')
    conn1.commit()
    conn1.close()

    conn2 = sqlite3.connect('equipment.db')
    c2 = conn2.cursor()
    c2.execute('''CREATE TABLE IF NOT EXISTS Equipment (
                 GipNum TEXT,
                 Type TEXT,
                 Model TEXT,
                 DateAdded TEXT,
                 SerialNum TEXT,
                 Hdetails TEXT,
                 Edetails TEXT,
                 DateUpdated TEXT,
                 State TEXT,
                 Owner TEXT,
                 DateAssumedReturn TEXT,
                 Tnum TEXT,
                 Location TEXT,
                 Notes TEXT,
                 WikiLink TEXT,
                 Returned TEXT
                 )''')
    conn2.commit()
    conn2.close()

# Function to clear fields in the loan tab
def clear_fields_loan():
    date_assumed_return_entry.delete(0, 'end')
    student_name_entry.delete(0, 'end')
    student_id_entry.delete(0, 'end')
    phone_num_entry.delete(0, 'end')
    email_entry.delete(0, 'end')
    project_tutor_entry.delete(0, 'end')
    gipnum_entry.delete(0, 'end')
    serialnum_entry.delete(0, 'end')

# Function to clear fields in the return tab
def clear_fields_return():
    gipnum_return_entry.delete(0, 'end')
    serialnum_return_entry.delete(0, 'end')


# Function to add a loan
def loan_item():
    date_loan = date_loaned_entry.get()
    date_assumed_return = date_assumed_return_entry.get()
    student_name = student_name_entry.get()
    student_id = student_id_entry.get()
    phone_num = phone_num_entry.get()
    email = email_entry.get()
    project_tutor = project_tutor_entry.get()
    gipnum = gipnum_entry.get()
    serialnum = serialnum_entry.get()

    if not (student_name or student_id or phone_num or email or project_tutor):
        messagebox.showwarning("Input Error", "Please fill in at least one field in Student Details.")
        return

    if not (gipnum or serialnum):
        messagebox.showwarning("Input Error", "Please provide GIP Number or Serial Number.")
        return

    add_loan(date_loan, student_name, student_id, phone_num, email, project_tutor, gipnum, serialnum, date_assumed_return)
    view_database()  # Refresh database view
    view_students()  # Refresh students view

# Modify add_loan function to store new fields
def add_loan(date_loan, student_name, student_id, phone_num, email, project_tutor, gipnum, serialnum, date_assumed_return):
    conn1 = sqlite3.connect('students.db')
    c1 = conn1.cursor()
    conn2 = sqlite3.connect('equipment.db')
    c2 = conn2.cursor()

    if gipnum:
        c2.execute('''SELECT * FROM Equipment WHERE GipNum = ?''', (gipnum,))
    else:
        c2.execute('''SELECT * FROM Equipment WHERE SerialNum = ?''', (serialnum,))

    record = c2.fetchone()
    if not record:
        messagebox.showwarning("Not Found", "No matching equipment record found.")
    else:
        if record[-1] == "NO":
            messagebox.showwarning("Not Available", "Equipment is already on loan.")
            return
        c1.execute('''INSERT INTO Students (DateLoan,DateAssumedReturn, StudentName, StudentID, PhoneNum, Email, ProjectTutor, GipNum, SerialNum)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                   (date_loan,date_assumed_return, student_name, student_id, phone_num, email, project_tutor, gipnum, serialnum))

        if gipnum:
            c2.execute('''UPDATE Equipment SET Returned = ?, DateAssumedReturn = ? WHERE GipNum = ?''', ("NO", date_assumed_return, gipnum))
        else:
            c2.execute('''UPDATE Equipment SET Returned = ?, DateAssumedReturn = ? WHERE SerialNum = ?''', ("NO", date_assumed_return, serialnum))

        messagebox.showinfo("Success", "Loan recorded successfully!")

    conn1.commit()
    conn2.commit()
    conn1.close()
    conn2.close()




# Function to return an item by GIP Number or Serial Number
def return_item(gipnum, serialnum):
    conn1 = sqlite3.connect('students.db')
    c1 = conn1.cursor()
    conn2 = sqlite3.connect('equipment.db')
    c2 = conn2.cursor()

    if gipnum:
        c1.execute('''DELETE FROM Students WHERE GipNum = ?''', (gipnum,))
        c2.execute('''UPDATE Equipment SET Returned = ? , DateAssumedReturn = ? WHERE GipNum = ?''', ("YES", "", gipnum))
    elif serialnum:
        c1.execute('''DELETE FROM Students WHERE SerialNum = ?''', (serialnum,))
        c2.execute('''UPDATE Equipment SET Returned = ?, DateAssumedReturn = ? WHERE SerialNum = ?''', ("YES","", serialnum))
    else:
        messagebox.showwarning("Input Error", "Please enter either GIP Number or Serial Number.")
        return

    conn1.commit()
    conn2.commit()
    conn1.close()
    conn2.close()
    messagebox.showinfo("Success", "Item returned successfully!")
    clear_fields_return()  # Clear the fields after returning
    view_database()  # This will refresh the Treeview with all the records, including the new one.
    view_students()  # This will refresh the Treeview with all the records, including the new one.


def return_loan():
    gipnum = gipnum_return_entry.get()
    serialnum = serialnum_return_entry.get()

    if not (gipnum or serialnum):
        messagebox.showwarning("Input Error", "Please enter either GIP Number or Serial Number.")
        return

    return_item(gipnum, serialnum)



def view_students():
    # Clear any previous widgets in the view_students_tab
    for widget in view_students_tab.winfo_children():
        widget.destroy()

    # Create the Treeview to display the student loan details
    tree = ttk.Treeview(view_students_tab, columns=("DateLoan", "DateAssumedReturn", "StudentName", "StudentID", "PhoneNum",
                                                     "Email", "ProjectTutor", "GipNum", "SerialNum"), show="headings", height=10)

    # Define column headings
    columns = ["DateLoan", "DateAssumedReturn", "StudentName", "StudentID", "PhoneNum", "Email", "ProjectTutor", "GipNum", "SerialNum"]

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=120)  # Adjust column width

    tree.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

    # Scrollbars
    vsb = ttk.Scrollbar(view_students_tab, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(view_students_tab, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    # Fetch data from the Students database
    conn = sqlite3.connect('students.db')
    c = conn.cursor()
    c.execute('SELECT * FROM Students')
    records = c.fetchall()

    for record in records:
        tree.insert("", "end", values=record)

    conn.close()

    # Define the filter function to show overdue students
    def filter_overdue_students():
        # Clear the current rows in the treeview before inserting the filtered data
        for row in tree.get_children():
            tree.delete(row)

        # Get today's date
        today = datetime.now().strftime('%Y-%m-%d')

        # Fetch overdue students from the database
        conn = sqlite3.connect('students.db')
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM Students WHERE DateAssumedReturn AND  DateAssumedReturn < ?', (today,))
        overdue_records = cursor.fetchall()

        # Insert the overdue records into the treeview
        for record in overdue_records:
            tree.insert("", "end", values=record)

        conn.close()

    # Define the reset function to reset the view to show all students
    def reset_view():
        view_students()

    # Create Filter Button to show overdue students
    filter_button = ttk.Button(view_students_tab, text="Show Overdue Students", command=filter_overdue_students)
    filter_button.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

    # Create Reset Button to reset view and show all students
    reset_button = ttk.Button(view_students_tab, text="Reset View", command=reset_view)
    reset_button.grid(row=3, column=0, padx=10, pady=10, sticky="ew")

# Function to view the database
def view_database():
    for widget in view_db_tab.winfo_children():
        widget.destroy()  # Clear previous widgets

    # Create Treeview
    tree = ttk.Treeview(view_db_tab, columns=("GipNum", "Type", "Model", "DateAdded",
                                              "SerialNum", "Hdetails", "Edetails", "DateUpdated", "State",
                                              "Owner", "DateAssumedReturn", "Tnum", "Location", "Notes",
                                              "WikiLink", "Returned"), show="headings", height=10)

    # Define column headings
    columns = ["GipNum", "Type", "Model", "DateAdded",
               "SerialNum", "Hdetails", "Edetails", "DateUpdated", "State",
               "Owner", "DateAssumedReturn", "Tnum", "Location", "Notes",
               "WikiLink", "Returned"]



    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=120)  # Adjust column width

    tree.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

    # Scrollbars
    vsb = ttk.Scrollbar(view_db_tab, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(view_db_tab, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    # Fetch data from database
    conn = sqlite3.connect('equipment.db')
    c = conn.cursor()
    c.execute('SELECT * FROM Equipment')
    records = c.fetchall()

    for record in records:
        tree.insert("", "end", values=record)

    conn.close()


    # Delete item from equipment and students database
    def delete_item():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Selection Error", "Please select an item to delete.")
            return

        item_values = tree.item(selected_item, 'values')
        gipnum = item_values[0]  # Assuming GipNum is in the 2nd column (index 1)
        serialnum = item_values[4]  # Assuming SerialNum is in the 6th column (index 5)

        # Check if GipNum is valid, otherwise use SerialNum
        if gipnum is None or gipnum == "":
            gipnum = serialnum  # Use SerialNum for deletion if GipNum is invalid

        # Delete from the database based on GipNum or SerialNum
        conn = sqlite3.connect('equipment.db')
        c = conn.cursor()

        # Try to delete by GipNum first
        if gipnum:
            c.execute("DELETE FROM Equipment WHERE GipNum = ?", (gipnum,))
        else:
            c.execute("DELETE FROM Equipment WHERE SerialNum = ?", (serialnum,))

        conn.commit()
        conn.close()

        # Delete from Students database
        conn1 = sqlite3.connect('students.db')
        c1 = conn1.cursor()
        if gipnum:
            c1.execute("DELETE FROM Students WHERE GipNum = ?", (gipnum,))
        else:
            c1.execute("DELETE FROM Students WHERE SerialNum = ?", (serialnum,))
        conn1.commit()
        conn1.close()

        # Remove from Treeview
        tree.delete(selected_item)
        messagebox.showinfo("Success", "Record deleted successfully!")
#edit item from equipment database
    def edit_item():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Selection Error", "Please select an item to edit.")
            return

        item_values = tree.item(selected_item, 'values')

        # Create a top-level window for editing
        edit_window = tk.Toplevel(root)
        edit_window.title("Edit Record")

        entry_widgets = {}
        for i, col in enumerate(columns):
            tk.Label(edit_window, text=col).grid(row=i, column=0, padx=5, pady=5, sticky="w")
            entry = tk.Entry(edit_window)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
            entry.insert(0, item_values[i])  # Pre-fill with current values

            if item_values[0] and i == 0:
                entry.config(state='disabled', disabledforeground='gray')
            if item_values[4] and i == 4:
                entry.config(state='disabled', disabledforeground='gray')
            if col == "Returned":
                entry.config(state='disabled', disabledforeground='gray')

            entry_widgets[col] = entry

        def save_changes():
            # Get all values, including from disabled fields
            new_values = [entry_widgets[col].get() for col in columns]
            gipnum = item_values[0]  # Original GipNum
            serialnum = item_values[4]  # Original SerialNum
            print(new_values)

            try:
                # Open the connection here (if not already open)
                conn = sqlite3.connect('equipment.db')
                c = conn.cursor()

                # Validate GipNum if provided
                if new_values[0]:
                    c.execute("SELECT COUNT(*) FROM Equipment WHERE GipNum = ? AND GipNum != ?",
                              (new_values[0], gipnum))
                    if c.fetchone()[0] > 1:  # Check for duplicate GipNum
                        messagebox.showwarning("Duplicate Error", f"GipNum {new_values[1]} already exists.")
                        return

                # Validate SerialNum if provided
                if new_values[4]:
                    c.execute("SELECT COUNT(*) FROM Equipment WHERE SerialNum = ? AND SerialNum != ?",
                              (new_values[4], serialnum))
                    if c.fetchone()[0] > 1:  # Check for duplicate SerialNum
                        messagebox.showwarning("Duplicate Error", f"SerialNum {new_values[5]} already exists.")
                        return

                # Perform the update, prioritizing GipNum
                if new_values[0]:
                    c.execute("""UPDATE Equipment SET 
                                    GipNum = ?, Type = ?, Model = ?, DateAdded = ?, SerialNum = ?, 
                                    Hdetails = ?, Edetails = ?, DateUpdated = ?, State = ?, Owner = ?, 
                                    DateAssumedReturn = ?, Tnum = ?, Location = ?, Notes = ?, WikiLink = ?, Returned = ? 
                                    WHERE GipNum = ?""", (
                        new_values[0], new_values[1], new_values[2], new_values[3], new_values[4],
                        new_values[5], new_values[6], new_values[7], new_values[8], new_values[9],
                        new_values[10], new_values[11], new_values[12], new_values[13], new_values[14],
                        new_values[15], gipnum  # Use original GipNum in WHERE clause
                    ))
                    conn.commit()

                elif new_values[4]:
                    c.execute("""UPDATE Equipment SET  
                                    GipNum = ?, Type = ?, Model = ?, DateAdded = ?, SerialNum = ?, 
                                    Hdetails = ?, Edetails = ?, DateUpdated = ?, State = ?, Owner = ?, 
                                    DateAssumedReturn = ?, Tnum = ?, Location = ?, Notes = ?, WikiLink = ?, Returned = ? 
                                    WHERE SerialNum = ?""", (
                        new_values[0], new_values[1], new_values[2], new_values[3], new_values[4],
                        new_values[5], new_values[6], new_values[7], new_values[8], new_values[9],
                        new_values[10], new_values[11], new_values[12], new_values[13], new_values[14],
                        new_values[15], serialnum
                    ))
                    conn.commit()

                else:
                    messagebox.showwarning("Input Error", "Please provide either GipNum or SerialNum.")
                    return

                tree.item(selected_item, values=new_values)  # Update Treeview
                messagebox.showinfo("Success", "Record updated successfully!")
                edit_window.destroy()

            except Exception as e:
                messagebox.showerror("Database Error", f"An error occurred: {str(e)}")
            finally:
                conn.close()  # Ensure the connection is closed after the operation

        tk.Button(edit_window, text="Save Changes", command=save_changes).grid(row=len(columns), columnspan=2, padx=10,
                                                                               pady=10)
# Function to add a new item to the equipment database
    def add_item():
        add_window = tk.Toplevel(root)
        add_window.title("Add New Record")

        entry_widgets = {}
        for i, col in enumerate(columns):
            tk.Label(add_window, text=col).grid(row=i, column=0, padx=5, pady=5, sticky="w")
            entry = tk.Entry(add_window)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
            entry_widgets[col] = entry

        def save_new_item():
            new_values = [entry_widgets[col].get() for col in columns]

            # Ensure 'Returned' field has a default value if empty
            new_values[15] = "YES"  # Default to "YES" for Returned if empty

            # Ensure that at least one field is filled in
            if all(value == "" for value in new_values):
                messagebox.showwarning("Input Error", "Please fill in at least one field.")
                return

            # Handle the two cases: GipNum or SerialNum
            gipnum = new_values[0]  # Assuming GipNum is the 2nd column (index 1)
            serialnum = new_values[4]  # Assuming SerialNum is the 6th column (index 5)

            if gipnum == "" and serialnum == "":
                messagebox.showwarning("Input Error", "Please fill in either GipNum or SerialNum.")
                return

            # Check if the GipNum or SerialNum already exists in the database
            conn = sqlite3.connect('equipment.db')
            c = conn.cursor()
            if gipnum:
                c.execute("SELECT COUNT(*) FROM Equipment WHERE GipNum = ?", (gipnum,))
                if c.fetchone()[0] > 0:
                    messagebox.showwarning("Duplicate Error",
                                           f"A record with GipNum: {gipnum} and SerialNum: {serialnum} already exists.")
                    conn.close()
                    return
            if serialnum:
                c.execute("SELECT COUNT(*) FROM Equipment WHERE SerialNum = ?", (serialnum,))
                if c.fetchone()[0] > 0:
                    messagebox.showwarning("Duplicate Error",
                                           f"A record with SerialNum: {serialnum} already exists.")
                    conn.close()
                    return

            # Insert into the database
            c.execute('''INSERT INTO Equipment (GipNum, Type, Model, DateAdded, SerialNum, Hdetails, 
                                                 Edetails, DateUpdated, State, Owner, DateAssumedReturn, 
                                                 Tnum, Location, Notes, WikiLink, Returned)
                         VALUES ( ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', new_values)
            conn.commit()
            conn.close()

            # Update the Treeview with the new record
            tree.insert("", "end", values=new_values)

            # Display success message
            messagebox.showinfo("Success", "Record added successfully!")

            # Close the 'Add New' window
            add_window.destroy()

            # Reload the database in the main view (optional if you need to refresh completely)
            view_database()  # This will refresh the Treeview with all the records, including the new one.

        tk.Button(add_window, text="Save Record", command=save_new_item).grid(row=len(columns), columnspan=2, padx=10,pady=10)

     # Function to export the database to an Excel file
    def export_to_xlsx():
        conn = sqlite3.connect('equipment.db')
        c = conn.cursor()
        c.execute("SELECT * FROM Equipment")
        records = c.fetchall()
        conn.close()

        # Create a new workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Equipment Data"

        # Add headers to the sheet
        sheet.append(["ID", "GipNum", "Type", "Model", "DateAdded", "SerialNum", "Hdetails",
                      "Edetails", "DateUpdated", "State", "Owner", "DateAssumedReturn",
                      "Tnum", "Location", "Notes", "WikiLink", "Returned"])

        # Add data rows to the sheet wile adding id col at beginning - a running number
        enumerated_records = enumerate(records, start=1)  # Start numbering from 1
        for idx, record in enumerated_records:
            # add id in first col and append the rest of the record
            recordee = (idx,) + record[0:]
            sheet.append(recordee)
        # Save the workbook to a file
        wb.save("equipment_data.xlsx")

        messagebox.showinfo("Success", "Database exported to Excel successfully!")

    # Function to import data from an Excel file
    def import_from_xlsx():
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

        if not file_path:
            return  # If the user cancels the file dialog, return

        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # Connect to the database
            conn = sqlite3.connect('equipment.db')
            c = conn.cursor()

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                gip_num = row[1]  # Assuming the GipNum is in the second column (index 1)
                serial_num = row[5]  # Assuming the SerialNum is in the sixth column (index 5)

                # Check if the GipNum or SerialNum already exists in the database
                if gip_num:
                    c.execute("SELECT COUNT(*) FROM Equipment WHERE GipNum = ?", (gip_num,))
                    result = c.fetchone()
                    if result and result[0] > 0:  # If the count is greater than 0, then the record already exists
                        messagebox.showwarning("Warning",
                                               f"Record with GipNum {gip_num} already exists. Skipping import.")
                        continue  # Skip this row if either GipNum or SerialNum already exists
                elif serial_num:
                    c.execute("SELECT COUNT(*) FROM Equipment WHERE SerialNum = ?", (serial_num,))
                    result = c.fetchone()
                    if result and result[0] > 0:  # If the count is greater than 0, then the record already exists
                        messagebox.showwarning("Warning",
                                               f"Record with SerialNum {serial_num} already exists. Skipping import.")
                        continue  # Skip this row if either GipNum or SerialNum already exists

                # Append "YES" for the "Returned" column
                row_with_returned = row[1:] + ("YES",)

                # Assuming the sheet columns are in a specific order, and adding "Returned" as the last column
                c.execute('''INSERT INTO Equipment (GipNum, Type, Model, DateAdded, SerialNum, 
                            Hdetails, Edetails, DateUpdated, State, Owner, DateAssumedReturn, Tnum, Location, Notes, WikiLink, Returned)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', row_with_returned)

            conn.commit()
            conn.close()

            messagebox.showinfo("Success", "Data imported successfully from Excel file!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while importing the Excel file: {str(e)}")

        view_database()  # This will refresh the Treeview with all the records, including the new one.

   # Adding buttons for export and import
    button_frame = tk.Frame(view_db_tab)
    button_frame.grid(row=4, column=0, columnspan=2, pady=10)

    tk.Button(button_frame, text="Export to Excel", command=export_to_xlsx).pack(side="left", padx=5)
    tk.Button(button_frame, text="Import from Excel", command=import_from_xlsx).pack(side="left", padx=5)

    # Function to filter the database based on selected option
    def option_selected(option):
        for row in tree.get_children():
            tree.delete(row)

        if option != "no filter":
            # Connect to your SQL database
            conn = sqlite3.connect('equipment.db')  # Replace with your database connection
            cursor = conn.cursor()

            # Query to select count of rows where Type is equal to the selected option
            query = "SELECT * FROM Equipment WHERE Type = ?"
            cursor.execute(query, (option,))

            recordsS = cursor.fetchall()

            for record in recordsS:
                tree.insert("", "end", values=record)

            conn.close()

        else:
            # If "no filter" is selected, show all records
            view_database()

        # Show the menu on button click
        def show_menu(event):
            menu.post(event.x_root, event.y_root)

        btn.bind("<Button-1>", show_menu)


    # Create a button
    btn = tk.Button(view_db_tab, text="Options")
    btn.grid(row=5, column=0, padx=5, pady=5)

    # Create a menu
    menu = tk.Menu(view_db_tab, tearoff=0)
    menu.add_command(label="Laptop", command=lambda: option_selected("Laptop"))
    menu.add_command(label="Monitor", command=lambda: option_selected("Monitor"))
    menu.add_command(label="PC", command=lambda: option_selected("PC"))
    menu.add_command(label="Camera", command=lambda: option_selected("Camera"))
    menu.add_command(label="Projector", command=lambda: option_selected("Projector"))
    menu.add_command(label="Hard Disk", command=lambda: option_selected("Hard Disk"))
    menu.add_command(label="Ptgray", command=lambda: option_selected("Ptgray"))
    menu.add_command(label="Lens", command=lambda: option_selected("Lens"))
    menu.add_command(label="Printer", command=lambda: option_selected("Printer"))
    menu.add_command(label="Tripod", command=lambda: option_selected("Tripod"))
    menu.add_command(label="Webcam", command=lambda: option_selected("Webcam"))
    menu.add_command(label="Micro Ball head", command=lambda: option_selected("Micro Ball head"))
    menu.add_command(label="Sensor", command=lambda: option_selected("Sensor"))
    menu.add_command(label="MAC", command=lambda: option_selected("MAC"))
    menu.add_command(label="Book", command=lambda: option_selected("Book"))
    menu.add_command(label="Oculus", command=lambda: option_selected("Oculus"))
    menu.add_command(label="Phone", command=lambda: option_selected("Phone"))
    menu.add_command(label="Tablet", command=lambda: option_selected("Tablet"))
    menu.add_command(label="Virtual reality", command=lambda: option_selected("Virtual reality"))

    menu.add_separator()
    menu.add_command(label="no filter", command=lambda: option_selected("no filter"))

    # Show the menu on button click
    def show_menu(event):
        menu.post(event.x_root, event.y_root)

    btn.bind("<Button-1>", show_menu)
    # Create a sub-frame to hold the buttons tightly together
    button_frame = tk.Frame(view_db_tab)
    button_frame.grid(row=3, column=0, columnspan=3, pady=10)

    # Add the buttons inside the frame, side by side
    tk.Button(button_frame, text="Add New", command=add_item).pack(side="left", padx=5)
    tk.Button(button_frame, text="Delete", command=delete_item).pack(side="left", padx=5)
    tk.Button(button_frame, text="Edit", command=edit_item).pack(side="left", padx=5)


# Setting up the main Tkinter window
root = tk.Tk()
root.title("Equipment Loan System")

# Allow window to resize
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

# Create the database if not exists
create_db()

# Create the notebook (tabs)
notebook = ttk.Notebook(root)
notebook.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

# Loan Tab
loan_tab = ttk.Frame(notebook)
notebook.add(loan_tab, text="Loan")

# Create frames for loan tab
loan_frame = ttk.Frame(loan_tab)
loan_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

# Loaned Item Additional Details Section
loan_details_frame = ttk.LabelFrame(loan_frame, text="Loan Details", padding=(10, 5))
loan_details_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

# Date Loaned
tk.Label(loan_details_frame, text="Date Loaned").grid(row=0, column=0, padx=5, pady=5, sticky="w")
date_loaned_entry = tk.Entry(loan_details_frame)
date_loaned_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

date_loaned_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))  # Auto-fill with current date

# Date Assumed Return
tk.Label(loan_details_frame, text="Date Assumed Return").grid(row=1, column=0, padx=5, pady=5, sticky="w")
date_assumed_return_entry = tk.Entry(loan_details_frame)
date_assumed_return_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")


# Student Details Section
student_frame = ttk.LabelFrame(loan_frame, text="Student Details", padding=(10, 5))
student_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

tk.Label(student_frame, text="Student Name").grid(row=0, column=0, padx=5, pady=5, sticky="w")
student_name_entry = tk.Entry(student_frame)
student_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

tk.Label(student_frame, text="Student ID").grid(row=1, column=0, padx=5, pady=5, sticky="w")
student_id_entry = tk.Entry(student_frame)
student_id_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

tk.Label(student_frame, text="Phone Number").grid(row=2, column=0, padx=5, pady=5, sticky="w")
phone_num_entry = tk.Entry(student_frame)
phone_num_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

tk.Label(student_frame, text="Email").grid(row=3, column=0, padx=5, pady=5, sticky="w")
email_entry = tk.Entry(student_frame)
email_entry.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

tk.Label(student_frame, text="Project Tutor").grid(row=4, column=0, padx=5, pady=5, sticky="w")
project_tutor_entry = tk.Entry(student_frame)
project_tutor_entry.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

# Loaned Item Details Section
item_frame = ttk.LabelFrame(loan_frame, text="Loaned Item Details", padding=(10, 5))
item_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

tk.Label(item_frame, text="GIP Number").grid(row=0, column=0, padx=5, pady=5, sticky="w")
gipnum_entry = tk.Entry(item_frame)
gipnum_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

tk.Label(item_frame, text="Serial Number").grid(row=1, column=0, padx=5, pady=5, sticky="w")
serialnum_entry = tk.Entry(item_frame)
serialnum_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

# Add Loan Button
loan_button = tk.Button(loan_frame, text="Record Loan", command=loan_item)
loan_button.grid(row=3, column=0, padx=10, pady=10)
#Add Reset Button
reset_loan_button = tk.Button(loan_tab, text="  Reset Fields  ", command=clear_fields_loan)
reset_loan_button.grid(row=3, column=0, padx=10, pady=10)




# Return Tab
return_tab = ttk.Frame(notebook)
notebook.add(return_tab, text="Return")

# Return Form
tk.Label(return_tab, text="GIP Number").grid(row=0, column=0, padx=5, pady=5, sticky="w")
gipnum_return_entry = tk.Entry(return_tab)
gipnum_return_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

tk.Label(return_tab, text="Serial Number").grid(row=1, column=0, padx=5, pady=5, sticky="w")
serialnum_return_entry = tk.Entry(return_tab)
serialnum_return_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

return_button = tk.Button(return_tab, text="Return Item", command=return_loan)
return_button.grid(row=2, columnspan=2, padx=10, pady=10)

# View Database Tab
view_db_tab = ttk.Frame(notebook)
notebook.add(view_db_tab, text="View Database")

# View Database Button
view_db_button = tk.Button(view_db_tab, text="Load Database", command=view_database)
view_db_button.grid(row=0, column=0, padx=10, pady=10)

# Configure the row/column of view_db_tab for resizing
view_db_tab.grid_rowconfigure(0, weight=1)
view_db_tab.grid_columnconfigure(0, weight=1)

# Add the View Students Tab to the Notebook
view_students_tab = ttk.Frame(notebook)
notebook.add(view_students_tab, text="View Students")
# Configure the row/column of view_db_tab for resizing
view_students_tab.grid_rowconfigure(0, weight=1)
view_students_tab.grid_columnconfigure(0, weight=1)
# Add a Button to the View Students Tab to allow viewing of students who have loaned equipment
tk.Button(view_students_tab, text="View Students", command=view_students).grid(row=0, column=0, padx=5, pady=5)

# Running the main window
root.mainloop()
